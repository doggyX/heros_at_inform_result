import os
import json
import time
import logging
import requests
import hashlib
import hmac
import base64
import urllib.parse
import uuid
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
from io import BytesIO

# 导入配置
import verify_config as vcfg

# 全局变量，用于存储任务处理状态
processed_tasks = set()

# 配置日志
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# 清除默认处理器
logger.handlers.clear()

# 创建控制台处理器
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler.setFormatter(console_formatter)
logger.addHandler(console_handler)

# 创建文件处理器
file_handler = logging.FileHandler(vcfg.VERIFY_CONFIG['log_file'], encoding='utf-8')
file_handler.setLevel(logging.INFO)
file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(filename)s:%(lineno)d - %(message)s')
file_handler.setFormatter(file_formatter)
logger.addHandler(file_handler)

class DingTalkRobot:
    """钉钉机器人客户端"""
    
    def __init__(self, webhook_url: str, secret: str):
        """
        初始化钉钉机器人
        
        Args:
            webhook_url: 钉钉机器人的Webhook地址
            secret: 加签密钥
        """
        self.webhook_url = webhook_url
        self.secret = secret
    
    def _generate_signature(self) -> Tuple[str, str]:
        """生成加签签名和时间戳"""
        timestamp = str(round(time.time() * 1000))
        secret_enc = self.secret.encode('utf-8')
        string_to_sign = f'{timestamp}\n{self.secret}'
        string_to_sign_enc = string_to_sign.encode('utf-8')
        hmac_code = hmac.new(secret_enc, string_to_sign_enc, 
                            digestmod=hashlib.sha256).digest()
        sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))
        return timestamp, sign
    
    def send_message(self, content: str = "", msgtype: str = "markdown", title: str = "Excel验证错误通知", btn_text: str = None, btn_url: str = None) -> bool:
        """
        发送消息到钉钉群
        
        Args:
            content: 消息内容
            msgtype: 消息类型，支持"text"、"markdown"或"actionCard"
            title: 消息标题，仅用于markdown类型
            btn_text: 按钮文本，仅用于actionCard类型
            btn_url: 按钮链接，仅用于actionCard类型
            
        Returns:
            发送是否成功
        """
        try:
            # 生成签名
            timestamp, sign = self._generate_signature()
            
            # 构建完整的webhook URL
            url = f"{self.webhook_url}&timestamp={timestamp}&sign={sign}"
            
            # 根据消息类型构建不同的消息结构
            if msgtype == "text":
                message = {
                    "msgtype": "text",
                    "text": {
                        "content": content
                    },
                    "at": {
                        "isAtAll": False
                    }
                }
            elif msgtype == "actionCard":
                # 构建actionCard消息
                message = {
                    "msgtype": "actionCard",
                    "actionCard": {
                        "title": title,
                        "text": content,
                        "btnOrientation": "0",
                        "btns": [
                            {
                                "title": btn_text or "已处理",
                                "actionURL": btn_url or ""
                            }
                        ]
                    }
                }
            else:  # markdown类型
                message = {
                    "msgtype": "markdown",
                    "markdown": {
                        "title": title,
                        "text": content
                    },
                    "at": {
                        "isAtAll": False
                    }
                }

            # 发送请求
            headers = {'Content-Type': 'application/json'}
            response = requests.post(url, data=json.dumps(message), headers=headers)
            
            if response.status_code == 200:
                result = response.json()
                if result.get('errcode') == 0:
                    logger.info(f"错误通知发送成功: {message}")
                    return True
                else:
                    logger.error(f"错误通知发送失败: {result}")
                    return False
            else:
                error_msg = f"HTTP请求失败: {response.status_code}"
                logger.error(error_msg)
                return False
                
        except Exception as e:
            error_msg = f"发送错误通知时发生错误: {str(e)}"
            logger.error(error_msg)
            return False

class ExcelVerifier:
    """Excel文件验证器"""
    
    def __init__(self, mount_config: Dict, task_file: str, error_robot: DingTalkRobot):
        """
        初始化Excel验证器
        
        Args:
            mount_config: 本地挂载配置
            task_file: 任务列表文件路径
            error_robot: 错误通知钉钉机器人实例
        """
        self.mount_config = mount_config
        self.task_file = task_file
        self.error_robot = error_robot
        self.tasks = self.load_tasks()
        self.command_file = 'control_command_verify.json'  # 验证程序的命令文件
        self.status_file = 'monitor_status_verify.json'  # 验证程序的状态文件
        self.last_command_time = 0  # 上次处理命令的时间
        self.check_interval = mount_config.get('check_interval', 1)  # 检查间隔（分钟）
        self.special_paths = self.load_special_paths()  # 加载特殊路径映射
        self.product_code_to_folder = self.special_paths.get('product_code_to_folder', {})
        self.product_code_to_cell = self.special_paths.get('product_code_to_cell', {})
        
        # 初始化状态文件
        self.update_status('running', '验证程序启动')
    
    def load_tasks(self) -> List[Dict]:
        """
        从文件加载任务列表
        
        Returns:
            任务列表
        """
        try:
            if os.path.exists(self.task_file):
                try:
                    with open(self.task_file, 'r', encoding='utf-8') as f:
                        content = f.read().strip()
                        if content:
                            tasks = json.loads(content)
                        else:
                            tasks = []
                    logger.info(f"从文件 {self.task_file} 加载了 {len(tasks)} 个任务")
                    return tasks
                except json.JSONDecodeError:
                    # 文件内容不是有效的JSON，创建空任务文件
                    with open(self.task_file, 'w', encoding='utf-8') as f:
                        json.dump([], f, ensure_ascii=False, indent=2)
                    logger.warning(f"任务文件格式无效，已重新创建空任务文件: {self.task_file}")
                    return []
            else:
                # 创建空任务文件
                with open(self.task_file, 'w', encoding='utf-8') as f:
                    json.dump([], f, ensure_ascii=False, indent=2)
                logger.info(f"创建空任务文件: {self.task_file}")
                return []
        except Exception as e:
            logger.error(f"加载任务文件失败: {str(e)}")
            return []
    
    def load_special_paths(self) -> Dict:
        """
        加载特殊路径映射文件
        
        Returns:
            特殊路径映射字典，包含product_code_to_folder和product_code_to_cell
        """
        try:
            special_path_file = vcfg.SPECIAL_PATH_CONFIG['file']
            if os.path.exists(special_path_file):
                with open(special_path_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                logger.info(f"从文件 {special_path_file} 加载了特殊路径映射")
                return {
                    'product_code_to_folder': data.get('product_code_to_folder', {}),
                    'product_code_to_cell': data.get('product_code_to_cell', {})
                }
            else:
                # 创建空的特殊路径映射文件
                with open(special_path_file, 'w', encoding='utf-8') as f:
                    json.dump({
                        'product_code_to_folder': {},
                        'product_code_to_cell': {}
                    }, f, ensure_ascii=False, indent=2)
                logger.info(f"创建空特殊路径映射文件: {special_path_file}")
                return {
                    'product_code_to_folder': {},
                    'product_code_to_cell': {}
                }
        except Exception as e:
            logger.error(f"加载特殊路径映射文件失败: {str(e)}")
            return {
                'product_code_to_folder': {},
                'product_code_to_cell': {}
            }
    
    def save_tasks(self, tasks: List[Dict]):
        """
        保存任务列表到文件
        
        Args:
            tasks: 任务列表
        """
        try:
            with open(self.task_file, 'w', encoding='utf-8') as f:
                json.dump(tasks, f, ensure_ascii=False, indent=2)
            logger.info(f"已保存 {len(tasks)} 个任务到文件: {self.task_file}")
        except Exception as e:
            logger.error(f"保存任务文件失败: {str(e)}")
    
    def update_status(self, status: str, message: str):
        """
        更新验证程序状态文件
        
        Args:
            status: 状态类型（running, paused, error, stopped）
            message: 状态描述
        """
        try:
            status_data = {
                'status': status,
                'message': message,
                'timestamp': time.time(),
                'check_interval': self.check_interval,
                'task_count': len(self.tasks)
            }
            
            with open(self.status_file, 'w', encoding='utf-8') as f:
                json.dump(status_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"更新状态文件失败: {str(e)}")
    
    def check_command(self):
        """
        检查并处理控制命令
        """
        try:
            if os.path.exists(self.command_file):
                # 获取文件修改时间
                file_mtime = os.path.getmtime(self.command_file)
                
                # 如果文件没有更新，跳过处理
                if file_mtime <= self.last_command_time:
                    return
                
                # 读取并解析命令
                with open(self.command_file, 'r', encoding='utf-8') as f:
                    command_data = json.load(f)
                
                # 更新上次处理命令时间
                self.last_command_time = file_mtime
                
                # 处理命令
                command = command_data.get('command')
                params = command_data.get('params', {})
                
                logger.info(f"收到命令: {command}，参数: {params}")
                
                if command == 'change_interval':
                    # 修改检查间隔
                    new_interval = params.get('interval')
                    if new_interval and isinstance(new_interval, (int, float)) and new_interval > 0:
                        self.check_interval = new_interval
                        self.update_status('running', f'检查间隔已修改为 {new_interval} 分钟')
                        logger.info(f"检查间隔已修改为 {new_interval} 分钟")
                    else:
                        logger.error("无效的检查间隔参数")
                elif command == 'clear_tasks':
                    # 清空任务列表
                    self.tasks = []
                    self.save_tasks([])
                    self.update_status('running', '任务列表已清空')
                    logger.info("任务列表已清空")
                
                # 删除已处理的命令文件
                os.remove(self.command_file)
        except Exception as e:
            logger.error(f"处理命令失败: {str(e)}")
    
    def test_mount_point(self) -> bool:
        """
        测试挂载点是否可用
        
        Returns:
            挂载点是否可用
        """
        try:
            mount_point = self.mount_config['mount_point']
            base_path = self.mount_config['base_path']
            
            # 检查挂载点是否存在
            if not os.path.exists(mount_point):
                logger.error(f"挂载点不存在: {mount_point}")
                return False
            
            # 检查基础路径是否存在
            full_path = os.path.join(mount_point, base_path)
            if not os.path.exists(full_path):
                logger.error(f"基础路径不存在: {full_path}")
                return False
            
            logger.info(f"挂载点测试成功: {mount_point}")
            return True
        except Exception as e:
            logger.error(f"挂载点测试失败: {str(e)}")
            return False
    
    def find_excel_file(self, product_code: str, batch_no: str, ipqc_no: str) -> Optional[str]:
        """
        在本地挂载路径中查找Excel文件
        
        Args:
            product_code: 产品编号
            batch_no: 生产批号
            ipqc_no: IPQC单号
            
        Returns:
            找到的文件路径，未找到返回None
        """
        try:
            mount_point = self.mount_config['mount_point']
            base_path = self.mount_config['base_path']
            
            # 先查找特殊路径映射
            folder_name = self.product_code_to_folder.get(product_code, product_code)
            
            # 构建产品编号文件夹路径
            product_folder = os.path.join(mount_point, base_path, folder_name)
            
            logger.info(f"查找产品文件夹: {product_folder}")
            
            # 检查产品文件夹是否存在
            if not os.path.exists(product_folder):
                logger.error(f"产品文件夹不存在: {product_folder}")
                return None
            
            # 递归查找文件
            return self._search_excel_recursive(product_folder, batch_no, ipqc_no)
        except Exception as e:
            logger.error(f"查找Excel文件失败: {str(e)}")
            return None
    
    def _search_excel_recursive(self, path: str, batch_no: str, ipqc_no: str) -> Optional[str]:
        """
        递归搜索Excel文件
        
        Args:
            path: 当前搜索路径
            batch_no: 生产批号
            ipqc_no: IPQC单号
            
        Returns:
            找到的文件路径，未找到返回None
        """
        try:
            for root, dirs, files in os.walk(path):
                for file in files:
                    # 检查是否是Excel文件
                    if file.lower().endswith(('.xlsx', '.xls')):
                        # 检查文件名是否同时包含生产批号和IPQC单号
                        if batch_no in file and ipqc_no in file:
                            full_path = os.path.join(root, file)
                            logger.info(f"找到匹配的Excel文件: {full_path}")
                            return full_path
            
            return None
        except Exception as e:
            logger.error(f"递归搜索失败: {str(e)}")
            return None
    
    def verify_excel_file(self, file_path: str, expected_result: str, product_code: str = '') -> Tuple[bool, str, dict]:
        """
        验证Excel文件
        
        Args:
            file_path: Excel文件路径
            expected_result: 预期结果
            product_code: 产品编号，用于查找预设的单元格位置
        
        Returns:
            (验证是否成功, 错误信息, 验证详情)
        """
        try:
            # 1. 检查文件是否存在
            if not os.path.exists(file_path):
                return False, f"文件不存在: {file_path}", {}
            
            # 2. 检查文件名中的pass/fail字段是否与结果相符
            file_name = os.path.basename(file_path)
            file_name_lower = file_name.lower()
            expected_result_lower = expected_result.lower()
            
            # 检查文件名是否包含预期结果或对应的值
            file_result_keyword = ""
            if expected_result_lower == 'pass':
                # pass可以对应OK
                if 'pass' in file_name_lower:
                    file_result_keyword = 'pass'
                elif 'ok' in file_name_lower:
                    file_result_keyword = 'ok'
                else:
                    return False, f"文件名中的结果与预期不符: {file_name}, 预期: {expected_result} (也可使用OK)", {}
            elif expected_result_lower == 'fail':
                # fail可以对应NG
                if 'fail' in file_name_lower:
                    file_result_keyword = 'fail'
                elif 'ng' in file_name_lower:
                    file_result_keyword = 'ng'
                else:
                    return False, f"文件名中的结果与预期不符: {file_name}, 预期: {expected_result} (也可使用NG)", {}
            else:
                # 其他结果直接检查
                if expected_result_lower in file_name_lower:
                    file_result_keyword = expected_result_lower
                else:
                    return False, f"文件名中的结果与预期不符: {file_name}, 预期: {expected_result}", {}
            
            # 3. 验证单元格的值
            try:
                import openpyxl
                
                # 加载Excel文件
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                
                # 获取第一张表
                if not wb.sheetnames:
                    return False, f"文件 {file_path} 没有工作表", {}
                
                sheet = wb[wb.sheetnames[0]]
                
                # 确定要验证的单元格
                cell_to_check = self.product_code_to_cell.get(product_code, 'F11')
                
                # 读取指定单元格的值
                try:
                    cell_value = sheet[cell_to_check].value
                except Exception as e:
                    return False, f"单元格 {cell_to_check} 不存在或无效: {str(e)}", {}
                
                # 检查单元格值是否与预期结果相符
                if cell_value is None:
                    return False, f"{cell_to_check}单元格为空: {file_path}", {}
                
                # 转换为字符串并小写
                cell_value_str = str(cell_value).lower()
                if expected_result_lower not in cell_value_str:
                    return False, f"{cell_to_check}单元格值与预期不符: {cell_value}, 预期: {expected_result}", {}
                
                wb.close()
                
                # 构建验证详情
                verify_details = {
                    'expected_result': expected_result,
                    'file_path': file_path,
                    'file_name': file_name,
                    'file_result_keyword': file_result_keyword,
                    'cell_checked': cell_to_check,
                    'cell_value': str(cell_value)
                }
                
                return True, "验证成功", verify_details
                
            except ImportError:
                return False, "openpyxl库未安装，无法验证Excel文件内容", {}
            except Exception as e:
                return False, f"读取Excel文件失败: {str(e)}", {}
                
        except Exception as e:
            return False, f"验证文件时发生错误: {str(e)}", {}
    
    def verify_task(self, task: Dict) -> Tuple[bool, str, float, dict]:
        """
        验证单个任务
        
        Args:
            task: 任务字典
            
        Returns:
            (验证是否成功, 错误信息, 耗时秒数, 验证详情)
        """
        start_time = time.time()
        
        try:
            product_code = (task.get('product_code') or '').strip()
            batch_no = (task.get('batch_no') or '').strip()
            ipqc_no = (task.get('ipqc_no') or '').strip()
            expected_result = (task.get('result') or '').strip()
            
            if not all([product_code, batch_no, ipqc_no, expected_result]):
                elapsed = time.time() - start_time
                return False, "任务信息不完整", elapsed, {}
            
            # 测试挂载点
            if not self.test_mount_point():
                elapsed = time.time() - start_time
                return False, "挂载点不可用", elapsed, {}
            
            # 查找Excel文件
            file_path = self.find_excel_file(product_code, batch_no, ipqc_no)
            
            if not file_path:
                elapsed = time.time() - start_time
                return False, f"未找到匹配的Excel文件: 产品编号={product_code}, 生产批号={batch_no}, IPQC单号={ipqc_no}", elapsed, {}
            
            # 验证Excel文件
            success, message, verify_details = self.verify_excel_file(file_path, expected_result, product_code)
            elapsed = time.time() - start_time
            
            return success, message, elapsed, verify_details
            
        except Exception as e:
            elapsed = time.time() - start_time
            return False, f"验证任务时发生错误: {str(e)}", elapsed, {}
    
    def send_error_notification(self, task: Dict, error_message: str, elapsed_time: float):
        """
        发送错误通知
        
        Args:
            task: 任务字典
            error_message: 错误消息
            elapsed_time: 验证耗时（秒）
        """
        try:
            # 构建错误通知消息
            title = "电气报告验证错误"
            # 使用markdown格式，确保错误信息可以换行
            content = f"# {title}\n\n"
            content += f"**错误时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
            content += f"**任务ID**: {task.get('task_id', '未知')}\n\n"
            content += f"**产品编号**: {task.get('product_code', '未知')}\n\n"
            content += f"**生产批号**: {task.get('batch_no', '未知')}\n\n"
            content += f"**IPQC单号**: {task.get('ipqc_no', '未知')}\n\n"
            content += f"**预期结果**: {task.get('result', '未知')}\n\n"
            content += f"**备注**: {task.get('remark', '无')}\n\n"
            content += f"**验证耗时**: {elapsed_time:.2f}秒\n\n"
            content += f"**错误信息**:\n{error_message}\n"
            
            # 生成任务处理URL（用于模拟回调）
            task_id = task.get('task_id', '')
            # 使用网络可访问的IP地址，而不是localhost
            # 这里使用127.0.0.1是示例，实际应使用电脑的局域网IP地址
            process_url = f"http://172.16.60.31:8090/process_task?task_id={task_id}"
            
            # 发送通知，使用actionCard类型添加已处理按钮
            self.error_robot.send_message(
                content=content, 
                msgtype="actionCard",
                title=title,
                btn_text="已处理，不再追踪",
                btn_url=process_url
            )
        except Exception as e:
            logger.error(f"发送错误通知失败: {str(e)}")
    
    def save_success_record(self, task: Dict, elapsed_time: float, verify_details: dict = {}):
        """
        保存验证成功的任务记录
        
        Args:
            task: 任务字典
            elapsed_time: 验证耗时（秒）
            verify_details: 验证详情
        """
        try:
            success_file = 'verify_success_records.json'
            success_records = []
            
            # 加载现有成功记录
            if os.path.exists(success_file):
                with open(success_file, 'r', encoding='utf-8') as f:
                    success_records = json.load(f)
            
            # 构建成功记录
            success_record = {
                'task_id': task.get('task_id'),
                'product_code': task.get('product_code'),
                'batch_no': task.get('batch_no'),
                'ipqc_no': task.get('ipqc_no'),
                'result': task.get('result'),
                'remark': task.get('remark', ''),
                'verified_at': datetime.now().isoformat(),
                'elapsed_time': elapsed_time,
                'status': 'success',
                'verify_details': {
                    'expected_result': verify_details.get('expected_result'),
                    'file_path': verify_details.get('file_path'),
                    'file_name': verify_details.get('file_name'),
                    'file_result_keyword': verify_details.get('file_result_keyword'),
                    'cell_checked': verify_details.get('cell_checked'),
                    'cell_value': verify_details.get('cell_value')
                }
            }
            
            # 添加到成功记录
            success_records.append(success_record)
            
            # 保存成功记录
            with open(success_file, 'w', encoding='utf-8') as f:
                json.dump(success_records, f, ensure_ascii=False, indent=2)
            
            logger.info(f"已保存成功记录: {task.get('task_id')}")
        except Exception as e:
            logger.error(f"保存成功记录失败: {str(e)}")
    
    def process_task(self, task_id: str):
        """
        处理用户点击'已处理，不再追踪'按钮的逻辑
        
        Args:
            task_id: 任务ID
        """
        try:
            # 加载最新任务列表
            self.tasks = self.load_tasks()
            
            # 查找任务
            task_to_process = None
            remaining_tasks = []
            
            for task in self.tasks:
                if task.get('task_id') == task_id:
                    task_to_process = task
                else:
                    remaining_tasks.append(task)
            
            if task_to_process:
                # 发送确认通知到群里
                title = "任务已处理通知"
                content = f"# {title}\n\n"
                content += f"**处理时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                content += f"**任务ID**: {task_to_process.get('task_id', '未知')}\n\n"
                content += f"**产品编号**: {task_to_process.get('product_code', '未知')}\n\n"
                content += f"**生产批号**: {task_to_process.get('batch_no', '未知')}\n\n"
                content += f"**IPQC单号**: {task_to_process.get('ipqc_no', '未知')}\n\n"
                content += f"**备注**: {task_to_process.get('remark', '无')}\n\n"
                content += f"**状态**: 已确认，不再跟踪\n"
                
                # 发送确认通知
                self.error_robot.send_message(content=content, msgtype="markdown", title=title)
                
                # 保存更新后的任务列表（移除已处理的任务）
                self.save_tasks(remaining_tasks)
                
                logger.info(f"任务 {task_id} 已处理，从任务列表中移除")
                return True
            else:
                logger.info(f"任务 {task_id} 不存在，可能已完成")
                return True
        except Exception as e:
            logger.error(f"处理任务时发生错误: {str(e)}")
            return False
    
    def process_tasks(self):
        """
        处理所有待验证任务
        """
        try:
            # 加载最新任务列表
            self.tasks = self.load_tasks()
            
            if not self.tasks:
                logger.info("没有待验证任务")
                return
            
            # 处理每个任务
            remaining_tasks = []
            
            for task in self.tasks:
                task_id = task.get('task_id')
                
                # 检查任务是否已被处理
                if task_id in processed_tasks:
                    logger.info(f"任务 {task_id} 已处理，跳过验证")
                    continue
                
                logger.info(f"开始验证任务: {task_id}")
                
                # 执行验证
                success, message, elapsed, verify_details = self.verify_task(task)
                
                if success:
                    logger.info(f"任务验证成功: {task_id}, 耗时: {elapsed:.2f}秒")
                    # 打关键结果日志
                    logger.info(f"验证详情: 预期结果={verify_details.get('expected_result')}, 查找到的文件={verify_details.get('file_path')}, 文件名关键字={verify_details.get('file_result_keyword')}, 检查单元格={verify_details.get('cell_checked')}, 单元格结果={verify_details.get('cell_value')}")
                    # 检查任务状态，如果是continue则发送复核成功通知
                    task_status = task.get('status', 'pending')
                    if task_status == 'continue':
                        # 发送复核成功通知
                        title = "任务复核成功通知"
                        content = f"# {title}\n\n"
                        content += f"**复核时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                        content += f"**任务ID**: {task_id}\n\n"
                        content += f"**产品编号**: {task.get('product_code', '未知')}\n\n"
                        content += f"**生产批号**: {task.get('batch_no', '未知')}\n\n"
                        content += f"**IPQC单号**: {task.get('ipqc_no', '未知')}\n\n"
                        content += f"**备注**: {task.get('remark', '无')}\n\n"
                        content += f"**状态**: 复核成功\n"
                        self.error_robot.send_message(content=content, msgtype="markdown", title=title)
                    # 保存成功记录
                    self.save_success_record(task, elapsed, verify_details)
                    # 验证成功，删除任务（不添加到remaining_tasks）
                else:
                    logger.error(f"任务验证失败: {task_id}, 错误: {message}, 耗时: {elapsed:.2f}秒")
                    # 发送错误通知
                    self.send_error_notification(task, message, elapsed)
                    # 验证失败，将任务状态改为continue
                    task['status'] = 'continue'
                    remaining_tasks.append(task)
            
            # 保存更新后的任务列表（只保留失败的任务）
            self.save_tasks(remaining_tasks)
            
            if remaining_tasks:
                logger.info(f"剩余 {len(remaining_tasks)} 个任务")
            else:
                logger.info("所有任务已完成")
        except Exception as e:
            logger.error(f"处理任务时发生错误: {str(e)}")
    
    def run(self, interval_minutes: int = 1):
        """
        运行验证程序
        
        Args:
            interval_minutes: 检查间隔（分钟）
        """
        self.check_interval = interval_minutes
        logger.info(f"开始运行Excel验证程序，间隔 {self.check_interval} 分钟")
        
        # 启动时测试挂载点
        logger.info("启动时测试挂载点...")
        self.test_mount_point()
        
        while True:
            try:
                # 检查控制命令
                self.check_command()
                
                # 处理任务
                self.process_tasks()
                
                # 等待下一次检查，期间定期检查命令
                wait_time = self.check_interval * 60
                elapsed = 0
                while elapsed < wait_time:
                    time.sleep(5)  # 每5秒检查一次命令
                    elapsed += 5
                    self.check_command()
                    # 检查间隔可能已更改，重新计算等待时间
                    new_wait_time = self.check_interval * 60
                    if new_wait_time != wait_time:
                        wait_time = new_wait_time
                        # 重置 elapsed，使用新的等待时间
                        elapsed = 0
                        logger.info(f"检查间隔已更改，新的等待时间: {wait_time}秒")
                
                logger.info(f"等待 {self.check_interval} 分钟后再次检查...")
            except KeyboardInterrupt:
                logger.info("验证程序被用户中断")
                break
            except Exception as e:
                logger.error(f"验证程序运行时发生错误: {str(e)}")
                time.sleep(60)  # 出错后等待1分钟再重试

class TaskManager:
    """任务管理器，用于添加验证任务"""
    
    def __init__(self, task_file: str, counter_file: str = 'task_counter.json'):
        """
        初始化任务管理器
        
        Args:
            task_file: 任务列表文件路径
            counter_file: 任务计数器配置文件路径
        """
        self.task_file = task_file
        self.counter_file = counter_file
        self.today_date = None  # 记录当天日期
        self.today_task_count = 0  # 记录当天任务数
        
        # 从配置文件加载任务计数器和日期
        self._load_counter()
    
    def _load_counter(self):
        """
        从配置文件加载任务计数器和日期
        """
        try:
            if os.path.exists(self.counter_file):
                with open(self.counter_file, 'r', encoding='utf-8') as f:
                    counter_data = json.load(f)
                self.today_date = counter_data.get('today_date')
                self.today_task_count = counter_data.get('today_task_count', 0)
                logger.info(f"从配置文件加载任务计数器: 日期={self.today_date}, 计数={self.today_task_count}")
            else:
                # 创建空的计数器文件
                self._save_counter()
                logger.info("创建新的任务计数器配置文件")
        except Exception as e:
            logger.error(f"加载任务计数器失败: {str(e)}")
            # 加载失败时初始化默认值
            self.today_date = None
            self.today_task_count = 0
    
    def _save_counter(self):
        """
        保存任务计数器和日期到配置文件
        """
        try:
            counter_data = {
                'today_date': self.today_date,
                'today_task_count': self.today_task_count
            }
            with open(self.counter_file, 'w', encoding='utf-8') as f:
                json.dump(counter_data, f, ensure_ascii=False, indent=2)
            logger.info(f"保存任务计数器: 日期={self.today_date}, 计数={self.today_task_count}")
        except Exception as e:
            logger.error(f"保存任务计数器失败: {str(e)}")
    
    def add_verify_task(self, task_data: Dict):
        """
        添加验证任务
        
        Args:
            task_data: 任务数据
        """
        try:
            # 加载现有任务
            tasks = []
            if os.path.exists(self.task_file):
                try:
                    with open(self.task_file, 'r', encoding='utf-8') as f:
                        content = f.read().strip()
                        if content:
                            tasks = json.loads(content)
                        else:
                            tasks = []
                except json.JSONDecodeError:
                    # 文件内容不是有效的JSON，使用空列表
                    tasks = []
            
            # 检查是否已存在相同的任务（通过产品编号、生产批号和IPQC单号判断）
            product_code = task_data.get('product_code')
            batch_no = task_data.get('batch_no')
            ipqc_no = task_data.get('ipqc_no')
            new_result = task_data.get('result')
            
            # 查找现有任务
            existing_task = None
            for task in tasks:
                if (task.get('product_code') == product_code and
                    task.get('batch_no') == batch_no and
                    task.get('ipqc_no') == ipqc_no):
                    existing_task = task
                    break
            
            if existing_task:
                # 任务已存在，修改预期结果
                existing_task['result'] = new_result
                # 更新备注字段
                existing_task['remark'] = task_data.get('remark', '')
                # 保持原来的状态，不重置为pending
                task_id = existing_task['task_id']
                logger.info(f"已更新验证任务: {task_id}, 预期结果改为: {new_result}")
            else:
                # 任务不存在，添加新任务
                # 生成任务ID
                # 从今天8点到次日8点日期为今天，加上3位编号
                now = datetime.now()
                # 检查是否在8点前，如果是则使用昨天的日期
                if now.hour < 8:
                    date = (now - timedelta(days=1)).strftime('%y%m%d')
                else:
                    date = now.strftime('%y%m%d')
                
                # 检查日期是否变化，如果是则重置任务计数
                if self.today_date != date:
                    self.today_date = date
                    # 初始化当天任务数
                    self.today_task_count = 0
                    # 遍历现有任务，统计当天的任务数（仅在日期变化时执行）
                    for task in tasks:
                        existing_task_id = task.get('task_id', '')
                        if existing_task_id.startswith(date):
                            self.today_task_count += 1
                    # 保存计数器配置
                    self._save_counter()
                
                # 递增任务数
                self.today_task_count += 1
                # 生成3位编号
                task_id = f"{date}{self.today_task_count:03d}"
                
                # 保存计数器配置
                self._save_counter()
                
                # 构建任务
                task = {
                    'task_id': task_id,
                    'product_code': product_code,
                    'batch_no': batch_no,
                    'ipqc_no': ipqc_no,
                    'result': new_result,
                    'remark': task_data.get('remark', ''),
                    'created_at': datetime.now().isoformat(),
                    'status': 'pending'
                }
                
                # 添加任务
                tasks.append(task)
                logger.info(f"已添加验证任务: {task_id}")
            
            # 保存任务
            with open(self.task_file, 'w', encoding='utf-8') as f:
                json.dump(tasks, f, ensure_ascii=False, indent=2)
            
            return task_id
        except Exception as e:
            logger.error(f"添加验证任务失败: {str(e)}")
            return None

def start_http_server(verifier):
    """
    启动HTTP服务器，用于处理钉钉按钮回调
    
    Args:
        verifier: ExcelVerifier实例
    """
    from http.server import HTTPServer, BaseHTTPRequestHandler
    import urllib.parse
    
    class RequestHandler(BaseHTTPRequestHandler):
        def do_GET(self):
                # 解析URL参数
                parsed_path = urllib.parse.urlparse(self.path)
                query_params = urllib.parse.parse_qs(parsed_path.query)
                
                # 检查是否是处理任务的请求
                if parsed_path.path == '/process_task':
                    task_id = query_params.get('task_id', [''])[0]
                    if task_id:
                        # 处理任务
                        success = verifier.process_task(task_id)
                        # 无论任务是否存在，都返回成功响应
                        # 标记任务为已处理
                        processed_tasks.add(task_id)
                        
                        # 返回成功响应（HTML页面）
                        self.send_response(200)
                        self.send_header('Content-type', 'text/html; charset=utf-8')
                        self.end_headers()
                        
                        if success:
                            # 任务存在且处理成功
                            html_content = '''
                            <!DOCTYPE html>
                            <html lang="zh-CN">
                            <head>
                                <meta charset="UTF-8">
                                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                                <title>任务处理成功</title>
                                <style>
                                    body {
                                        font-family: Arial, sans-serif;
                                        text-align: center;
                                        padding: 50px;
                                        background-color: #f5f5f5;
                                    }
                                    .container {
                                        background-color: white;
                                        padding: 30px;
                                        border-radius: 8px;
                                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                                        max-width: 500px;
                                        margin: 0 auto;
                                    }
                                    .success {
                                        color: #4CAF50;
                                        font-size: 18px;
                                        margin-bottom: 20px;
                                    }
                                    .message {
                                        color: #333;
                                        margin-bottom: 30px;
                                    }
                                    .btn {
                                        background-color: #4CAF50;
                                        color: white;
                                        padding: 10px 20px;
                                        border: none;
                                        border-radius: 4px;
                                        cursor: pointer;
                                        font-size: 14px;
                                    }
                                    .btn:hover {
                                        background-color: #45a049;
                                    }
                                </style>
                            </head>
                            <body>
                                <div class="container">
                                    <h2 class="success">任务处理成功</h2>
                                    <p class="message">任务 ''' + task_id + ''' 已确认，不再跟踪</p>
                                    <button class="btn" onclick="window.close()">关闭</button>
                                </div>
                            </body>
                            </html>
                            '''
                        else:
                            # 任务不存在，显示任务已完成
                            html_content = '''
                            <!DOCTYPE html>
                            <html lang="zh-CN">
                            <head>
                                <meta charset="UTF-8">
                                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                                <title>任务已完成</title>
                                <style>
                                    body {
                                        font-family: Arial, sans-serif;
                                        text-align: center;
                                        padding: 50px;
                                        background-color: #f5f5f5;
                                    }
                                    .container {
                                        background-color: white;
                                        padding: 30px;
                                        border-radius: 8px;
                                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                                        max-width: 500px;
                                        margin: 0 auto;
                                    }
                                    .success {
                                        color: #4CAF50;
                                        font-size: 18px;
                                        margin-bottom: 20px;
                                    }
                                    .message {
                                        color: #333;
                                        margin-bottom: 30px;
                                    }
                                    .btn {
                                        background-color: #4CAF50;
                                        color: white;
                                        padding: 10px 20px;
                                        border: none;
                                        border-radius: 4px;
                                        cursor: pointer;
                                        font-size: 14px;
                                    }
                                    .btn:hover {
                                        background-color: #45a049;
                                    }
                                </style>
                            </head>
                            <body>
                                <div class="container">
                                    <h2 class="success">任务已完成</h2>
                                    <p class="message">任务 ''' + task_id + ''' 已完成，无需处理</p>
                                    <button class="btn" onclick="window.close()">关闭</button>
                                </div>
                            </body>
                            </html>
                            '''
                        self.wfile.write(html_content.encode('utf-8'))
                    else:
                        # 缺少任务ID（HTML页面）
                        self.send_response(400)
                        self.send_header('Content-type', 'text/html; charset=utf-8')
                        self.end_headers()
                        html_content = """
                        <!DOCTYPE html>
                        <html lang="zh-CN">
                        <head>
                            <meta charset="UTF-8">
                            <meta name="viewport" content="width=device-width, initial-scale=1.0">
                            <title>参数错误</title>
                            <style>
                                body {
                                    font-family: Arial, sans-serif;
                                    text-align: center;
                                    padding: 50px;
                                    background-color: #f5f5f5;
                                }
                                .container {
                                    background-color: white;
                                    padding: 30px;
                                    border-radius: 8px;
                                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                                    max-width: 500px;
                                    margin: 0 auto;
                                }
                                .error {
                                    color: #f44336;
                                    font-size: 18px;
                                    margin-bottom: 20px;
                                }
                                .message {
                                    color: #333;
                                    margin-bottom: 30px;
                                }
                                .btn {
                                    background-color: #f44336;
                                    color: white;
                                    padding: 10px 20px;
                                    border: none;
                                    border-radius: 4px;
                                    cursor: pointer;
                                    font-size: 14px;
                                }
                                .btn:hover {
                                    background-color: #da190b;
                                }
                            </style>
                        </head>
                        <body>
                            <div class="container">
                                <h2 class="error">参数错误</h2>
                                <p class="message">缺少任务ID参数</p>
                                <button class="btn" onclick="window.close()">关闭</button>
                            </div>
                        </body>
                        </html>
                        """
                        self.wfile.write(html_content.encode('utf-8'))
                else:
                    # 其他路径（HTML页面）
                    self.send_response(404)
                    self.send_header('Content-type', 'text/html; charset=utf-8')
                    self.end_headers()
                    html_content = """
                    <!DOCTYPE html>
                    <html lang="zh-CN">
                    <head>
                        <meta charset="UTF-8">
                        <meta name="viewport" content="width=device-width, initial-scale=1.0">
                        <title>页面不存在</title>
                        <style>
                            body {
                                font-family: Arial, sans-serif;
                                text-align: center;
                                padding: 50px;
                                background-color: #f5f5f5;
                            }
                            .container {
                                background-color: white;
                                padding: 30px;
                                border-radius: 8px;
                                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                                max-width: 500px;
                                margin: 0 auto;
                            }
                            .error {
                                color: #f44336;
                                font-size: 18px;
                                margin-bottom: 20px;
                            }
                            .message {
                                color: #333;
                                margin-bottom: 30px;
                            }
                            .btn {
                                background-color: #666;
                                color: white;
                                padding: 10px 20px;
                                border: none;
                                border-radius: 4px;
                                cursor: pointer;
                                font-size: 14px;
                            }
                            .btn:hover {
                                background-color: #555;
                            }
                        </style>
                    </head>
                    <body>
                        <div class="container">
                            <h2 class="error">页面不存在</h2>
                            <p class="message">请求的页面不存在</p>
                            <button class="btn" onclick="window.close()">关闭</button>
                        </div>
                    </body>
                    </html>
                    """
                    self.wfile.write(html_content.encode('utf-8'))
    
    # 创建服务器
    server_address = ('', 8090)
    httpd = HTTPServer(server_address, RequestHandler)
    logger.info("HTTP服务器已启动，监听端口 8090")
    
    # 启动服务器
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        logger.info("HTTP服务器已停止")
        httpd.shutdown()


def main():
    """主函数"""
    # 创建错误通知钉钉机器人
    error_robot = DingTalkRobot(
        webhook_url=vcfg.ERROR_DINGTALK_CONFIG['webhook_url'],
        secret=vcfg.ERROR_DINGTALK_CONFIG['secret']
    )
    
    # 创建Excel验证器
    verifier = ExcelVerifier(
        mount_config=vcfg.MOUNT_CONFIG,
        task_file=vcfg.VERIFY_CONFIG['task_file'],
        error_robot=error_robot
    )
    
    # 启动HTTP服务器线程
    import threading
    http_thread = threading.Thread(target=start_http_server, args=(verifier,), daemon=True)
    http_thread.start()
    
    # 运行验证程序
    verifier.run(interval_minutes=vcfg.VERIFY_CONFIG['check_interval'])

if __name__ == "__main__":
    main()
